import psycopg2
import pandas as pd
from datetime import datetime, timedelta
import time
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, Alignment, NamedStyle, PatternFill
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv
import os


def generate_xls(file_path):
    if not os.path.exists("xlsx"):
        os.mkdir("xlsx")

    df = pd.DataFrame(rows, columns=columns)
    df.to_excel(file_path, index=False)

    wb = load_workbook(file_path)
    ws = wb.active

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="000000", end_color="000000", fill_type="solid"
    )

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col_num, col_name in enumerate(df.columns, start=1):
        col_letter = get_column_letter(col_num)
        ws[f"{col_letter}1"].font = header_font
        ws[f"{col_letter}1"].fill = header_fill
        ws[f"{col_letter}1"].border = thin_border
        ws[f"{col_letter}1"].alignment = Alignment(
            horizontal="center", vertical="center"
        )

    number_format = NamedStyle(name="number_format")
    number_format.number_format = "#,##0"

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if isinstance(cell.value, (int, float)):
                cell.number_format = number_format.number_format

    for col_num, column_cells in enumerate(ws.columns, start=1):
        max_length = 0
        col_letter = get_column_letter(col_num)
        for cell in column_cells:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    wb.save(file_path)


conn = None
load_dotenv()
try:
    start_time = int(round(time.time() * 1000))

    print("Conectando no banco ao banco de dados ...")
    conn = psycopg2.connect(
        dbname=os.getenv("DB_NAME"),
        user=os.getenv("DB_USER"),
        password=os.getenv("DB_PASSWORD"),
        host=os.getenv("DB_HOST"),
        port=os.getenv("DB_PORT"),
    )
    print("Conexão efetuada com sucesso!")

    date_interval = int(input("Digite n° de intervalo de dias:"))
    print("Gerando planilha de métricas...")
    num_cycles = 3

    timezone_offset = timedelta(hours=3)
    today = datetime.utcnow() + timezone_offset
    today = today.replace(hour=0, minute=0, second=0, microsecond=0)

    cycles = []
    start_date = today - timedelta(days=num_cycles * date_interval)
    for i in range(num_cycles):
        end_date = start_date + timedelta(
            days=date_interval - 1, hours=23, minutes=59, seconds=59
        )
        cycles.append((start_date, end_date))
        start_date = end_date + timedelta(seconds=1)

    count_clauses = []
    for start, end in cycles:
        clause = (
            f'COUNT(CASE WHEN sv."createdAt" BETWEEN \'{start.strftime("%Y-%m-%d 00:00:00")}\' '
            f'AND \'{end.strftime("%Y-%m-%d 23:59:59")}\' THEN 1 END) AS "Quant_Erros_{start.strftime("%d/%m")} a {end.strftime("%d/%m")}"'
        )

        count_clauses.append(clause)

    with conn.cursor() as cursor:

        sql_query = f"""
            SELECT 
                sv.error_code as "Código De Erro",
                CASE 
                    WHEN sv.error_message ~ 'API_EMAIL_' THEN 'API_EMAIL_*_CAN_NOT_BE_CREATED_TICKET_LINKED_TO_ANOTHER_PROVIDER'
                    WHEN sv.error_message ~ 'SYNC_ERROR_EXCEPTION_' THEN 'SYNC_ERROR_EXCEPTION_*'
                    WHEN sv.error_message ~ 'JÁ POSSUI UM TICKET BÁSICO' THEN 'EP0001: JÁ POSSUI UM TICKET BÁSICO ATIVO: *'
                    WHEN sv.error_message ~ 'EMAIL_FROM_' THEN 'P0001: WARNING: [EMAIL_FROM_*_IS_BLOCKED]'
                    ELSE sv.error_message 
                END AS "Descrição do Erro",
                {', '.join(count_clauses)}
            FROM 
                sva_errors sv
            WHERE 
                sv."createdAt" BETWEEN '{cycles[0][0].strftime("%Y-%m-%d %H:%M:%S")}' AND '{cycles[-1][1].strftime("%Y-%m-%d %H:%M:%S")}'
                AND sv.error_code > 0
                AND sv.error_message IS NOT NULL
                AND sv.error_message != ''
            GROUP BY 
                "Descrição do Erro", 
                sv.error_code
            ORDER BY         
                5 DESC,
                4 DESC,
                3 DESC;
        """

        cursor.execute(sql_query)

        rows = cursor.fetchall()
        columns = [desc[0] for desc in cursor.description]

        file_path = f"xlsx/METRICAS_HUBSVA{int(round(time.time() * 1000))}.xlsx"

        generate_xls(file_path)
        end_time = int(round(time.time() * 1000))

        print(
            f"Planilha de métricas gerada no arquivo {file_path} \n {end_time - start_time}ms"
        )


except Exception as e:
    print(e)

finally:
    if conn is not None:
        conn.close()
